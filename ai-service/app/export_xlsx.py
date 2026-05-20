from io import BytesIO
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook


def _pick_str(data: Dict[str, Any], *keys: str, default: str = "") -> str:
    if not isinstance(data, dict):
        return default
    lowered = {str(k).lower(): v for k, v in data.items()}
    for key in keys:
        value = data.get(key)
        if value is None:
            value = lowered.get(str(key).lower())
        if value is None:
            continue
        if isinstance(value, list):
            parts = [str(item).strip() for item in value if str(item).strip()]
            if parts:
                return "; ".join(parts)
        text = str(value).strip()
        if text:
            return text
    return default


def _pick_int(data: Dict[str, Any], *keys: str, default: Union[int, str] = "") -> Union[int, str]:
    if not isinstance(data, dict):
        return default
    lowered = {str(k).lower(): v for k, v in data.items()}
    for key in keys:
        value = data.get(key)
        if value is None:
            value = lowered.get(str(key).lower())
        if value is None:
            continue
        try:
            return int(value)
        except (TypeError, ValueError):
            continue
    return default


def _normalize_requirement_row(item: Any) -> Dict[str, Any]:
    if not isinstance(item, dict):
        return {
            "id": str(item).strip() if item is not None else "",
            "feature": "",
            "inputFields": "",
            "conditions": "",
            "expectedAction": "",
        }
    inputs = item.get("inputFields") or item.get("inputs") or item.get("input") or item.get("input_fields")
    if isinstance(inputs, list):
        input_text = ", ".join(str(part).strip() for part in inputs if str(part).strip())
    else:
        input_text = str(inputs or "").strip()

    conditions = item.get("conditions") or item.get("condition")
    if isinstance(conditions, list):
        cond_text = "; ".join(str(part).strip() for part in conditions if str(part).strip())
    else:
        cond_text = str(conditions or "").strip()

    return {
        "id": _pick_str(item, "id", "reqId", "requirementId", "requirement_id"),
        "feature": _pick_str(item, "feature", "name", "title", "module", "component", "description"),
        "inputFields": input_text,
        "conditions": cond_text,
        "expectedAction": _pick_str(
            item,
            "expectedAction",
            "expected",
            "expectedResult",
            "expected_result",
            "action",
            "description",
        ),
    }


def _normalize_risk_row(item: Any) -> Dict[str, Any]:
    if not isinstance(item, dict):
        return {
            "reqId": "",
            "impact": "",
            "likelihood": "",
            "riskScore": "",
            "priority": "",
            "rationale": str(item or ""),
        }
    return {
        "reqId": _pick_str(item, "reqId", "req_id", "id", "requirementId", "requirement_id"),
        "impact": _pick_int(item, "impact", "Impact"),
        "likelihood": _pick_int(item, "likelihood", "Likelihood"),
        "riskScore": _pick_int(item, "riskScore", "risk_score", "score", "RiskScore"),
        "priority": _pick_str(item, "priority", "Priority", "level"),
        "rationale": _pick_str(item, "rationale", "reason", "description", "note"),
    }


def _normalize_strategy_row(item: Any) -> Dict[str, Any]:
    if not isinstance(item, dict):
        return {"id": "", "method": "", "name": "", "isoRef": "", "linkedTestcases": str(item or "")}
    linked = item.get("linkedTestcases") or item.get("linked_testcases") or item.get("testcases") or []
    if isinstance(linked, list):
        linked_text = ", ".join(str(part).strip() for part in linked if str(part).strip())
    else:
        linked_text = str(linked or "").strip()
    return {
        "id": _pick_str(item, "id", "strategyId", "strategy_id"),
        "method": _pick_str(item, "method", "designMethod", "technique"),
        "name": _pick_str(item, "name", "title", "strategyName"),
        "isoRef": _pick_str(item, "isoRef", "iso_ref", "standard", "iso"),
        "linkedTestcases": linked_text,
    }


def _normalize_testcase_row(item: Any) -> Dict[str, Any]:
    if not isinstance(item, dict):
        return {
            "id": "",
            "designMethod": "",
            "title": "",
            "priority": "",
            "oracle": "",
            "expected": "",
            "steps": "",
        }
    steps = item.get("steps") or item.get("procedure") or item.get("actions")
    if isinstance(steps, list):
        steps_text = " | ".join(str(part).strip() for part in steps if str(part).strip())
    else:
        steps_text = str(steps or "").strip()

    return {
        "id": _pick_str(item, "id", "testCaseId", "testcaseId", "caseId", "tcId"),
        "designMethod": _pick_str(item, "designMethod", "design_method", "method", "technique"),
        "title": _pick_str(item, "title", "name", "summary", "description"),
        "priority": _pick_str(item, "priority", "Priority", "severity"),
        "oracle": _pick_str(item, "oracle", "testOracle", "expectedOracle"),
        "expected": _pick_str(item, "expected", "expectedResult", "expected_result", "outcome"),
        "steps": steps_text,
    }


def _coerce_artifacts(artifacts: Any) -> Dict[str, Any]:
    if not isinstance(artifacts, dict):
        return {}
    merged = dict(artifacts)
    for wrapper in ("data", "artifacts", "artifact", "result", "output"):
        nested = merged.get(wrapper)
        if isinstance(nested, dict):
            for key, value in nested.items():
                if key not in merged or merged[key] in (None, [], {}):
                    merged[key] = value
    return merged


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def build_xlsx_bytes(artifacts: Dict[str, Any], testcases: List[Dict[str, Any]]) -> bytes:
    artifacts = _coerce_artifacts(artifacts)
    requirements = artifacts.get("requirementsStructured") or artifacts.get("structuredRequirements") or []
    risks = artifacts.get("riskItems") or artifacts.get("risks") or artifacts.get("riskAnalysis") or []
    strategies = artifacts.get("testStrategies") or artifacts.get("strategies") or []
    cases = testcases or artifacts.get("testcases") or artifacts.get("testCases") or []

    if not isinstance(requirements, list):
        requirements = []
    if not isinstance(risks, list):
        risks = []
    if not isinstance(strategies, list):
        strategies = []
    if not isinstance(cases, list):
        cases = []

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Requirements"
    _write_sheet(
        ws1,
        ["id", "feature", "inputFields", "conditions", "expectedAction"],
        [_normalize_requirement_row(r) for r in requirements],
    )

    ws2 = wb.create_sheet("Risks")
    _write_sheet(
        ws2,
        ["reqId", "impact", "likelihood", "riskScore", "priority", "rationale"],
        [_normalize_risk_row(r) for r in risks],
    )

    ws3 = wb.create_sheet("Strategies")
    _write_sheet(
        ws3,
        ["id", "method", "name", "isoRef", "linkedTestcases"],
        [_normalize_strategy_row(s) for s in strategies],
    )

    ws4 = wb.create_sheet("TestCases")
    _write_sheet(
        ws4,
        ["id", "designMethod", "title", "priority", "oracle", "expected", "steps"],
        [_normalize_testcase_row(c) for c in cases],
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
