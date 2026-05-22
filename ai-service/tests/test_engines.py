import unittest
import asyncio
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.export_xlsx import build_xlsx_bytes
from app.engines.blackbox_fallbacks import generate_blackbox_artifacts, _pairwise_combinations
from app.engines.generation_pipeline import GlobalContext, run_generation_pipeline
from app.engines.risk_engine import compute_risk_score, score_requirements
from app.engines.schema_validator import validate_llm_payload
from app.engines.state_model_engine import parse_custom_state_model
from app.engines.whitebox_coverage import generate_coverage_items
from app.engines.whitebox_java_analyzer import analyze_java_source
from app.engines.whitebox_java_worker import generate_java_whitebox_design, worker_whitebox_java, _sequences_to_testcases
from app.engines.whitebox_llm_enhancer import (
    build_llm_enhancement_prompt,
    enhance_whitebox_design_with_llm,
    parse_llm_enhancement_response,
)
from app.engines.whitebox_sequence_generator import generate_test_sequences

try:
    import javalang  # noqa: F401
    HAS_JAVALANG = True
except Exception:
    HAS_JAVALANG = False


LOGIN_JAVA = """
public class LoginService {
    public String login(String username, String password) {
        if (username == null || password == null) {
            return "missing";
        }
        if (username.equals("admin") && password.equals("123456")) {
            return "success";
        }
        return "invalid";
    }
}
"""


def _coverage_items(design, item_type):
    return [item for item in design["coverageItems"] if item.get("type") == item_type]


def _sequence_for_target(design, target_id):
    for sequence in design["testSequences"]:
        if target_id in sequence.get("coverageTargets", []):
            return sequence
    return None


def _path_text(sequence):
    return "\n".join(str(step.get("text") or step.get("label") or "") for step in sequence.get("path", []))


class EngineTests(unittest.TestCase):
    def test_risk_score_formula(self):
        self.assertEqual(compute_risk_score(5, 4), 20)

    def test_csv_pipeline(self):
        content = """[CSV requirements]
id,feature,input,condition,expected
REQ-REC-001,记录过滤,count+duration,count<3 and duration<30,filtered
"""
        requirements = [
            {
                "id": "REQ-REC-001",
                "feature": "record filtering",
                "inputFields": ["count", "durationSeconds"],
                "conditions": ["count < 3", "durationSeconds < 30"],
                "expectedAction": "filtered",
            }
        ]
        result = run_generation_pipeline(
            GlobalContext.from_pipeline_kwargs(
                requirements=requirements,
                risk_items=score_requirements(requirements),
                whitebox_description="UP -> DOWN -> UP",
            ),
            ["EP", "BVA", "DecisionTable"],
            include_oracle=False,
            include_optimization=False,
        )
        self.assertGreaterEqual(len(result["requirementsStructured"]), 1)
        self.assertGreaterEqual(len(result["testcases"]), 3)
        methods = {c["designMethod"] for c in result["testcases"]}
        self.assertTrue({"EP", "BVA", "DecisionTable"}.issubset(methods))

    def test_whitebox_custom_model(self):
        model = parse_custom_state_model('{"states":["A","B"],"transitions":[{"from":"A","to":"B"}]}')
        self.assertEqual(model["states"], ["A", "B"])

    def test_schema_validator(self):
        payload = {
            "requirementsStructured": [{"id": "R1"}],
            "riskItems": [{"reqId": "R1", "impact": 2, "likelihood": 2, "riskScore": 4}],
            "testcases": [
                {"id": "1", "designMethod": "EP"},
                {"id": "2", "designMethod": "BVA"},
                {"id": "3", "designMethod": "DecisionTable"},
            ],
        }
        ok, issues = validate_llm_payload(payload)
        self.assertTrue(ok)
        self.assertEqual(issues, [])

    def test_pairwise(self):
        rows = _pairwise_combinations([["A", "B"], ["X", "Y"]], ["f1", "f2"], max_rows=4)
        self.assertGreaterEqual(len(rows), 2)

    def test_export_xlsx_field_aliases(self):
        artifacts = {
            "requirementsStructured": [
                {
                    "id": "REQ-1",
                    "description": "姿态分析",
                    "inputs": ["exerciseType"],
                    "condition": "valid",
                    "expected": "返回 ok",
                }
            ],
            "riskItems": [
                {"id": "REQ-1", "impact": 5, "likelihood": 4, "risk_score": 20, "priority": "high"}
            ],
            "testStrategies": [
                {
                    "strategyId": "S1",
                    "technique": "EP",
                    "title": "EP strategy",
                    "standard": "ISO 29119-4",
                    "testcases": ["TC-1"],
                }
            ],
        }
        cases = [
            {
                "testCaseId": "TC-1",
                "method": "EP",
                "name": "valid partition",
                "severity": "high",
                "expected_result": "ok",
                "steps": ["step1", "step2"],
            }
        ]
        workbook = load_workbook(BytesIO(build_xlsx_bytes(artifacts, cases)), read_only=True)
        req = workbook["Requirements"]
        risk = workbook["Risks"]
        case = workbook["TestCases"]
        self.assertEqual(req.cell(2, 1).value, "REQ-1")
        self.assertEqual(req.cell(2, 2).value, "姿态分析")
        self.assertEqual(req.cell(2, 3).value, "exerciseType")
        self.assertEqual(risk.cell(2, 1).value, "REQ-1")
        self.assertEqual(risk.cell(2, 4).value, 20)
        self.assertEqual(case.cell(2, 1).value, "TC-1")
        self.assertEqual(case.cell(2, 2).value, "EP")
        self.assertEqual(case.cell(2, 6).value, "ok")

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box analysis")
    def test_whitebox_java_login_analysis_and_coverage(self):
        analysis = analyze_java_source(LOGIN_JAVA, "LoginService.java")
        self.assertEqual(analysis["warnings"], [])
        methods = analysis["classes"][0]["methods"]
        self.assertEqual(methods[0]["name"], "login")
        self.assertEqual(len(methods[0]["parameters"]), 2)
        self.assertIn("cfg", methods[0])
        self.assertTrue(methods[0]["cfg"]["nodes"])
        self.assertTrue(methods[0]["cfg"]["edges"])
        self.assertGreaterEqual(len([s for s in methods[0]["statements"] if s["kind"] == "return"]), 3)
        self.assertEqual(len(methods[0]["decisions"]), 2)

        coverage, warnings = generate_coverage_items(analysis, "statement+branch")
        self.assertEqual(warnings, [])
        self.assertTrue(any(item["type"] == "statement" for item in coverage))
        branch_items = [item for item in coverage if item["type"] == "branch"]
        self.assertGreaterEqual(len(branch_items), 4)
        self.assertTrue(all(item.get("sourceEdgeId") for item in branch_items))
        self.assertTrue(any(item["branchValue"] is True for item in branch_items))
        self.assertTrue(any(item["branchValue"] is False for item in branch_items))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_sequences_and_reviewer_overrides(self):
        analysis = analyze_java_source(LOGIN_JAVA, "LoginService.java")
        coverage, _warnings = generate_coverage_items(
            analysis,
            "branch",
            {
                "coverageItemSelection": {"COV-BR-001-F": False},
                "manualCoverageItems": [
                    {
                        "id": "COV-MANUAL-001",
                        "type": "manual",
                        "methodId": "M-001",
                        "target": "Cover invalid password retry scenario",
                        "selected": True,
                    }
                ],
            },
        )
        sequences = generate_test_sequences(analysis, coverage)
        targets = {target for seq in sequences for target in seq["coverageTargets"]}
        self.assertNotIn("COV-BR-001-F", targets)
        self.assertIn("COV-MANUAL-001", targets)
        branch_sequences = [seq for seq in sequences if any(str(target).startswith("COV-BR") for target in seq["coverageTargets"])]
        self.assertTrue(branch_sequences)
        self.assertTrue(any(step["type"] == "edge" and step.get("edgeId") for seq in branch_sequences for step in seq["path"]))
        self.assertTrue(any(step["type"] == "node" and step.get("nodeType") == "decision" for seq in branch_sequences for step in seq["path"]))
        self.assertTrue(any(seq["inputHints"] for seq in branch_sequences))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_sequential_statement_path_merges(self):
        source = """
public class Seq {
    public int calc(int x) {
        int y = x + 1;
        y = y * 2;
        return y;
    }
}
"""
        design = generate_java_whitebox_design(source, "Seq.java", "statement")
        statement_items = _coverage_items(design, "statement")
        self.assertGreaterEqual(len(statement_items), 3)
        self.assertEqual(len(design["testSequences"]), 1)
        targets = set(design["testSequences"][0]["coverageTargets"])
        self.assertTrue({item["id"] for item in statement_items}.issubset(targets))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_if_early_return_false_path_uses_cfg(self):
        source = """
public class Early {
    public String check(String name) {
        if (name == null) {
            return "missing";
        }
        return "present";
    }
}
"""
        design = generate_java_whitebox_design(source, "Early.java", "branch")
        false_item = next(item for item in _coverage_items(design, "branch") if item.get("branchLabel") == "false")
        sequence = _sequence_for_target(design, false_item["id"])
        self.assertIsNotNone(sequence)
        text = _path_text(sequence)
        self.assertIn('return "present";', text)
        self.assertNotIn('return "missing";', text)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_if_else_reaches_distinct_returns(self):
        source = """
public class Branchy {
    public String sign(int value) {
        if (value >= 0) {
            return "positive";
        } else {
            return "negative";
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "Branchy.java", "branch")
        true_item = next(item for item in _coverage_items(design, "branch") if item.get("branchLabel") == "true")
        false_item = next(item for item in _coverage_items(design, "branch") if item.get("branchLabel") == "false")
        true_text = _path_text(_sequence_for_target(design, true_item["id"]))
        false_text = _path_text(_sequence_for_target(design, false_item["id"]))
        self.assertIn('return "positive";', true_text)
        self.assertNotIn('return "negative";', true_text)
        self.assertIn('return "negative";', false_text)
        self.assertNotIn('return "positive";', false_text)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_nested_if_paths_remain_reachable(self):
        source = """
public class Nested {
    public String grade(int score) {
        if (score >= 60) {
            if (score >= 90) {
                return "high";
            }
            return "pass";
        }
        return "fail";
    }
}
"""
        design = generate_java_whitebox_design(source, "Nested.java", "branch")
        outer_false = next(item for item in _coverage_items(design, "branch") if item.get("branchLabel") == "false" and "score >= 60" in item.get("target", ""))
        sequence = _sequence_for_target(design, outer_false["id"])
        self.assertIsNotNone(sequence)
        text = _path_text(sequence)
        self.assertIn('return "fail";', text)
        self.assertNotIn('return "high";', text)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_loop_has_enter_and_skip_branches(self):
        source = """
public class Looping {
    public int sum(int n) {
        int total = 0;
        while (n > 0) {
            total = total + n;
            n--;
        }
        return total;
    }
}
"""
        design = generate_java_whitebox_design(source, "Looping.java", "branch")
        labels = {item.get("branchLabel") for item in _coverage_items(design, "branch")}
        self.assertIn("true", labels)
        self.assertIn("false", labels)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_switch_generates_case_and_default_branches(self):
        source = """
public class Switcher {
    public String map(int code) {
        switch (code) {
            case 1:
                return "one";
            case 2:
                return "two";
            default:
                return "other";
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "Switcher.java", "branch")
        labels = {str(item.get("branchLabel")) for item in _coverage_items(design, "branch")}
        self.assertTrue(any("1" in label for label in labels))
        self.assertTrue(any("2" in label for label in labels))
        self.assertIn("default", labels)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_try_catch_generates_normal_and_exception_branches(self):
        source = """
public class Parser {
    public int parse(String raw) {
        try {
            return Integer.parseInt(raw);
        } catch (NumberFormatException ex) {
            return -1;
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "Parser.java", "branch")
        labels = {str(item.get("branchLabel")) for item in _coverage_items(design, "branch")}
        self.assertIn("normal", labels)
        self.assertTrue(any(label.startswith("exception") for label in labels))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_try_conditions_and_exception_hints_are_specific(self):
        source = """
public class Parser {
    public int parse(String raw) {
        try {
            return Integer.parseInt(raw);
        } catch (NumberFormatException ex) {
            return -1;
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "Parser.java", "branch")
        method = design["analysis"]["classes"][0]["methods"][0]
        try_decision = next(decision for decision in method["decisions"] if decision["kind"] == "try")
        self.assertEqual(try_decision["condition"], "try/catch control")
        edges = method["cfg"]["edges"]
        normal_edge = next(edge for edge in edges if edge["type"] == "normal" and edge["from"] == try_decision["sourceNodeId"])
        exception_edge = next(edge for edge in edges if edge["type"] == "exception")
        self.assertEqual(normal_edge["condition"], "try block completes without a caught exception")
        self.assertEqual(exception_edge["condition"], "a caught exception is thrown in the try block")
        self.assertFalse(any("EEPTION" in item["id"] for item in _coverage_items(design, "branch")))
        exception_item = next(item for item in _coverage_items(design, "branch") if str(item.get("branchLabel")).startswith("exception"))
        self.assertTrue(exception_item["id"].endswith("-EXC"))
        self.assertNotIn("EXCEPTIO", exception_item["id"])
        exception_sequence = _sequence_for_target(design, exception_item["id"])
        self.assertTrue(exception_sequence["exceptionTriggerHints"])
        self.assertIn("parseInt", exception_sequence["exceptionTriggerHints"][0]["text"])

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_multiple_catches_use_exc_indexes(self):
        source = """
public class MultiCatch {
    public int parse(String raw) {
        try {
            return Integer.parseInt(raw);
        } catch (NumberFormatException ex) {
            return -1;
        } catch (RuntimeException ex) {
            return -2;
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "MultiCatch.java", "branch")
        ids = [item["id"] for item in _coverage_items(design, "branch") if str(item.get("branchLabel")).startswith("exception")]
        self.assertTrue(any(item.endswith("-EXC1") for item in ids))
        self.assertTrue(any(item.endswith("-EXC2") for item in ids))
        self.assertFalse(any("EXCEPTIO" in item or "EEPTION" in item for item in ids))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_switch_fallthrough_reaches_next_case(self):
        source = """
public class SwitchFallthrough {
    public String map(int code) {
        switch (code) {
            case 1:
                code++;
            case 2:
                return "two";
            default:
                return "other";
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "SwitchFallthrough.java", "branch")
        case_one = next(item for item in _coverage_items(design, "branch") if str(item.get("branchLabel")) == "1")
        sequence = _sequence_for_target(design, case_one["id"])
        text = _path_text(sequence)
        self.assertIn("code++;", text)
        self.assertIn('return "two";', text)
        self.assertNotIn('return "other";', text)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_hints_separate_inputs_setups_and_conflicts(self):
        source = """
public class Hinty {
    public String run(int x, String raw) {
        int local = Integer.parseInt(raw);
        if (x > 0) {
            if (x < 0) {
                return "impossible";
            }
        }
        if (local == 1) {
            return "local";
        }
        return "done";
    }
}
"""
        design = generate_java_whitebox_design(source, "Hinty.java", "branch")
        impossible_item = next(item for item in _coverage_items(design, "branch") if "x < 0" in item.get("target", "") and item.get("branchLabel") == "true")
        impossible_sequence = _sequence_for_target(design, impossible_item["id"])
        self.assertIn("x", impossible_sequence["inputHints"])
        self.assertTrue(impossible_sequence["constraintConflicts"])
        local_item = next(item for item in _coverage_items(design, "branch") if "local == 1" in item.get("target", "") and item.get("branchLabel") == "true")
        local_sequence = _sequence_for_target(design, local_item["id"])
        self.assertNotIn("local", local_sequence["inputHints"])
        self.assertTrue(any("local" in hint for hint in local_sequence["setupHints"]))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_oracle_hints_http_statuses(self):
        source = """
public class Controller {
    public Object ok() { return ResponseEntity.ok("done"); }
    public Object bad() { return ResponseEntity.badRequest().body("bad"); }
    public Object fail() { return ResponseEntity.internalServerError().body("fail"); }
}
"""
        design = generate_java_whitebox_design(source, "Controller.java", "statement")
        statuses = {seq["oracleHints"]["httpStatusHint"] for seq in design["testSequences"]}
        self.assertIn("200", statuses)
        self.assertIn("400", statuses)
        self.assertIn("500", statuses)

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_multiline_response_return_is_readable(self):
        source = """
public class Controller {
    public Object ok() {
        return ResponseEntity.ok(Map.of(
            "status", "ok",
            "count", 1
        ));
    }
}
"""
        design = generate_java_whitebox_design(source, "Controller.java", "statement")
        oracle = design["testSequences"][0]["oracleHints"]
        self.assertEqual(oracle["httpStatusHint"], "200")
        self.assertIn("ResponseEntity.ok", oracle["returnText"])
        self.assertIn("Map.of", oracle["returnText"])
        self.assertFalse(oracle["returnText"].endswith("("))
        self.assertFalse(oracle["bodyHint"].endswith("("))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_testcase_exposes_llm_ready_fields(self):
        source = """
public class Parser {
    public int parse(String raw) {
        try {
            return Integer.parseInt(raw);
        } catch (NumberFormatException ex) {
            return -1;
        }
    }
}
"""
        design = generate_java_whitebox_design(source, "Parser.java", "branch")
        cases = _sequences_to_testcases(design["testSequences"], design["analysis"], 1)
        exception_case = next(case for case in cases if case["exceptionTriggerHints"])
        self.assertIn("pathConstraints", exception_case)
        self.assertIn("setupHints", exception_case)
        self.assertIn("constraintConflicts", exception_case)
        self.assertIn("oracleHints", exception_case)
        self.assertIn("needsReview", exception_case)
        self.assertIn("coverageTargets", exception_case)
        self.assertIn("Configure one of the possible throw sites", exception_case["setup"])

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_exception_without_trigger_needs_review(self):
        source = """
public class EmptyTry {
    public int run() {
        try {
            int value = 1;
        } catch (RuntimeException ex) {
            return -1;
        }
        return 0;
    }
}
"""
        design = generate_java_whitebox_design(source, "EmptyTry.java", "branch")
        exception_item = next(item for item in _coverage_items(design, "branch") if str(item.get("branchLabel")).startswith("exception"))
        sequence = _sequence_for_target(design, exception_item["id"])
        self.assertEqual(sequence["exceptionTriggerHints"], [])
        self.assertTrue(sequence["needsReview"])
        self.assertTrue(any("No concrete throw site" in hint for hint in sequence["setupHints"]))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_titles_are_specific(self):
        source = """
public class Titles {
    public String run(int x, int y) {
        if (x > 0) {
            return "x";
        }
        if (y > 0) {
            return "y";
        }
        return "none";
    }
}
"""
        design = generate_java_whitebox_design(source, "Titles.java", "branch")
        false_titles = [seq["title"] for seq in design["testSequences"] if "== false" in seq["title"]]
        self.assertEqual(len(false_titles), len(set(false_titles)))
        self.assertTrue(any("x > 0 == false" in title for title in false_titles))
        self.assertTrue(any("y > 0 == false" in title for title in false_titles))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box sequence generation")
    def test_whitebox_java_input_hints_exclude_internal_variables(self):
        source = """
public class InternalVar {
    public String run(String raw) {
        int local = Integer.parseInt(raw);
        if (local == 1) {
            return "one";
        }
        return "other";
    }
}
"""
        design = generate_java_whitebox_design(source, "InternalVar.java", "branch")
        local_item = next(item for item in _coverage_items(design, "branch") if "local == 1" in item.get("target", ""))
        sequence = _sequence_for_target(design, local_item["id"])
        self.assertNotIn("local", sequence["inputHints"])
        self.assertTrue(sequence["setupHints"])

    def test_whitebox_java_worker_metadata_declares_llm_boundary_without_javalang(self):
        ctx = SimpleNamespace(
            extra={"sourceContent": "public class Broken {"},
            whitebox_description="",
            coverage_criterion="statement+branch",
        )
        result = asyncio.run(worker_whitebox_java(ctx, 1))
        metadata = result.artifacts["engineMetadata"]
        self.assertEqual(metadata["identificationMode"], "deterministic CFG-based coverage item identification")
        self.assertIn("must not create/delete CFG coverage items", metadata["llmBoundary"])
        self.assertIn("llmReadyWhiteboxContext", result.artifacts)
        self.assertIn("llmEnhancedTestcases", result.artifacts)

    def test_whitebox_llm_enhancer_parse_ignores_forbidden_fields(self):
        parsed, warnings = parse_llm_enhancement_response(
            {
                "naturalLanguageTitle": "Reject invalid pose analysis request",
                "testIntentSummary": "Verify invalid request is rejected.",
                "refinedInputSuggestions": ["Send request.isValid() == false equivalent input."],
                "coverageItems": [{"id": "COV-HACK"}],
                "path": [{"nodeId": "N-HACK"}],
            }
        )
        self.assertEqual(parsed["naturalLanguageTitle"], "Reject invalid pose analysis request")
        self.assertEqual(parsed["testIntentSummary"], "Verify invalid request is rejected.")
        self.assertEqual(parsed["refinedInputSuggestions"], ["Send request.isValid() == false equivalent input."])
        self.assertTrue(any("Ignored forbidden LLM fields" in warning for warning in warnings))

    def test_whitebox_llm_enhancer_prompt_preview_without_client(self):
        design = {
            "sourceFile": "Demo.java",
            "sourceSnippet": "public class Demo {}",
            "coverageItems": [{"id": "COV-BR-001-T", "type": "branch"}],
            "testSequences": [
                {
                    "id": "SEQ-WB-001",
                    "methodId": "M-001",
                    "title": "Cover demo branch",
                    "coverageTargets": ["COV-BR-001-T"],
                    "path": [{"type": "edge", "edgeId": "E-001", "label": "true", "condition": "x > 0"}],
                    "pathConstraints": ["x > 0 == true"],
                    "setupHints": [],
                    "exceptionTriggerHints": [],
                    "oracleHints": {"httpStatusHint": "unknown"},
                    "needsReview": True,
                }
            ],
            "testcases": [{"sequenceId": "SEQ-WB-001", "title": "Cover demo branch"}],
        }
        enhanced = enhance_whitebox_design_with_llm(design, llm_client=None)
        item = enhanced["llmEnhancedTestcases"][0]
        self.assertEqual(item["baseSequenceId"], "SEQ-WB-001")
        self.assertIn("Deterministic context", item["promptPreview"])
        self.assertIn("do not create, delete, rename, or modify CFG", item["promptPreview"])
        self.assertEqual(item["naturalLanguageTitle"], "")

    def test_whitebox_llm_enhancer_without_real_client_uses_prompt_preview(self):
        design = {
            "sourceFile": "Demo.java",
            "coverageItems": [],
            "testSequences": [{"id": "SEQ-WB-001", "coverageTargets": [], "path": []}],
        }
        enhanced = enhance_whitebox_design_with_llm(design, llm_client=None)
        item = enhanced["llmEnhancedTestcases"][0]
        self.assertEqual(item["naturalLanguageTitle"], "")
        self.assertTrue(item["promptPreview"])
        self.assertTrue(any("Real LLM client is not configured" in warning for warning in item["reviewerWarnings"]))

    @unittest.skipUnless(HAS_JAVALANG, "javalang is required for Java white-box analysis")
    def test_whitebox_java_nested_class_constructor_is_analyzed(self):
        source = """
public class Outer {
    class Inner {
        Inner(String name) {
            if (name == null) {
                throw new IllegalArgumentException();
            }
        }
    }
}
"""
        analysis = analyze_java_source(source, "Outer.java")
        names = {class_item["name"] for class_item in analysis["classes"]}
        self.assertIn("Outer.Inner", names)
        inner = next(class_item for class_item in analysis["classes"] if class_item["name"] == "Outer.Inner")
        self.assertEqual(inner["methods"][0]["kind"], "constructor")

    def test_whitebox_java_invalid_source_returns_warning(self):
        analysis = analyze_java_source("public class Broken {", "Broken.java")
        self.assertEqual(analysis["classes"], [])
        self.assertTrue(analysis["warnings"])


if __name__ == "__main__":
    unittest.main()
